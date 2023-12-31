# Required modules for regex patterns, dataframes, turtle graphics, and excel files
import re
import pandas
import turtle
from openpyxl.styles import Font, Color, PatternFill, Border, Side
from tabulate import tabulate
from turtle import Screen, Turtle
from openpyxl import Workbook, load_workbook


# turtle graphics screen
def start():
    # Set up the turtle graphics window
    turtle.bgcolor('black')
    height = 500
    width = 1000
    screen = Screen()
    screen.title("Price Comparison Calculator by Harveer Talwar")
    screen.setup(width, height)
    t = turtle.Turtle()
    t.color('white')
    t.shape('turtle')

    # Move the turtle to the starting point
    t.penup()
    t.goto(-100, 0)
    t.pendown()

    # Draw the 'S' shape and horizontal lines
    t.forward(45)
    t.circle(50, 180)
    t.circle(-50, 180)
    t.forward(45)

    # Draw the first horizontal line
    t.penup()
    t.goto(-70, -10)
    t.setheading(90)
    t.pendown()
    t.forward(225)

    # Draw the second horizontal line
    t.penup()
    t.goto(-50, -10)
    t.setheading(90)
    t.pendown()
    t.forward(225)
    t.penup()

    # Display welcome message and instructions
    t.goto(-50, -60)
    style = "Comic Sans MS", 25, "normal"
    second_style = "Comic Sans MS", 15, "normal"
    t.write("Welcome to Price Comparison Calculator", font=style, align='center')
    t.goto(-50, -100)
    t.pendown()
    t.hideturtle()
    t.write("Close this window to begin using the calculator", font=second_style, align='center')
    turtle.done()


# Define the instructions function to display program instructions
def instructions():
    print("""***** Instructions *****
    Enter the name of an item, quantity (e.g., 120kg, 10l), and its total cost
    To finish entering items, type 'xxx' when prompted for the item name
    After entering all items, you will see a table with price information
    The program will identify the best option within your budget
    All the price comparison information will be added to an auto-generated Excel file
    You can view the Excel file at any time and share it with people of your choice
    Enjoy using the calculator!
    """)


# validating quantity and unit input
def validate_quantity_unit(user_input, error):
    while True:
        response = input(user_input).lower()
        # Regular expression pattern to match the quantity and unit
        pattern = r'^(0*[1-9]\d*(\.\d+)?|0*\.\d*[1-9]\d*)(kg|g|ml|l)$'

        # Compile the pattern into a regular expression object
        regex = re.compile(pattern, re.IGNORECASE)

        # Use re.match() to check if the user input matches the pattern
        match = regex.match(response)

        if match:
            # If a match is found, extract the quantity and unit from the input
            quantity = float(match.group(1))  # Convert the quantity to a float
            unit = match.group(3).lower()  # Get the unit and convert it to lowercase

            return [quantity, unit, response]
        else:
            # If no match is found, return None
            print(error)
            continue


# ormatting a number as currency
def currency(x):
    return f"${x:.2f}"


# yes/no questions
def yes_no(question):
    to_check = ["yes", "no"]
    while True:
        response = input(question).lower()
        for var_item in to_check:
            if response == var_item:
                return response
            elif response == var_item[0]:
                return var_item
        print("Please enter either yes or no... ..\n")


# checking non-numeric strings
def string_check(question, error):
    while True:
        response = input(question)
        if response.isnumeric():
            print(f"{error}. \nPlease try again. \n")
            continue

        return response


# checking and returning numeric values
def num_check(question, error, num_type):
    while True:
        try:
            response = num_type(input(question))
            if response <= 0:
                print(error)
                continue
            return float(response)
        except ValueError:
            print(error)


# gather item details and perform price comparison
def get_items():
    per_unit = None
    item_cost = None
    quantity = None
    user_budget = num_check("What is your budget: $", error="Please enter a number more than 0", num_type=float)
    item_list = []
    quantity_list = []
    converted_quantity_list = []
    cost_list = []
    per_unit_list = []
    item_dict = {
        "Item": item_list,
        "Amount": quantity_list,
        "Converted amount": converted_quantity_list,
        "Cost": cost_list,
        "Unit Price": per_unit_list
    }

    while True:
        item_name = string_check("Item name: ", "The item name cannot be blank or a number")
        if item_name == "xxx":
            break
        quantity = validate_quantity_unit("What is the quantity (e.g 120kg, 10l): ",
                                          error="Please enter a valid quantity")
        item_cost = num_check("What is the total cost: $", error="Please enter a valid cost", num_type=float)
        conversion_dict = {
            "l": 1 * quantity[0],
            "ml": 0.001 * quantity[0],
            "g": 0.001 * quantity[0],
            "kg": 1 * quantity[0]
        }
        converted_quantity = conversion_dict[quantity[1]]
        unit_cost = round((item_cost / converted_quantity), 2)
        if quantity[1] == "ml" or quantity[1] == "l":
            unit = "L"
        else:
            unit = "KG"
        per_unit = f"${unit_cost}/{unit}"
        item_list.append(item_name)
        quantity_list.append(quantity[2])
        converted_quantity_list.append(f"{converted_quantity}{unit}")
        cost_list.append(item_cost)
        per_unit_list.append(per_unit)

    price_frame = pandas.DataFrame(item_dict)
    price_frame = price_frame.set_index('Item')
    price_frame[['Cost']] = price_frame[['Cost']].applymap(currency)
    price_frame = price_frame.sort_values(by='Unit Price', ascending=True)
    table = tabulate(price_frame, headers='keys', tablefmt='fancy_grid')
    print(table)

    price_frame['Unit Price Numeric'] = price_frame['Unit Price'].str.replace(r'\$|/.*', '', regex=True).astype(float)
    price_frame['Cost'] = price_frame['Cost'].str.replace('$', '').astype(float)
    affordable_items = price_frame[price_frame['Cost'] <= float(user_budget)]
    affordable_items = affordable_items.sort_values(by='Unit Price Numeric')
    price_frame = price_frame.drop('Unit Price Numeric', axis=1)

    if not affordable_items.empty:
        best_option = affordable_items.iloc[0]
        best_option_name = best_option.name
        conclusion = f"The best option within your budget (${user_budget}) is: {best_option_name}, {per_unit}, {quantity[2]} costs ${item_cost}"
    else:
        conclusion = "There are no affordable options"
        best_option_name = "no affordable options"
    print(conclusion)
    budget_info = f"Budget: ${user_budget}"
    price_frame[['Cost']] = price_frame[['Cost']].applymap(currency)
    file_name = f'{best_option_name}.xlsx'
    price_frame.to_excel(file_name)
    wb = load_workbook(file_name)
    ws = wb.active
    ws.insert_rows(1)
    ws.insert_rows(2)
    ws.insert_rows(3)
    ws.merge_cells("A1:E1")
    ws.merge_cells("A2:E2")
    ws["A1"] = budget_info
    ws["A2"] = conclusion
    fill = PatternFill(fgColor="FFC7CE", fill_type="solid")
    fill2 = PatternFill(fgColor="0000FF", fill_type="solid")
    ws["A1"].fill = fill
    ws["A2"].fill = fill2
    wb.save(file_name)
    return [table, price_frame, user_budget]


# for displaying the main menu
def show_menu():
    print("""\n Menu
    1. View Instructions 📝
    2. Start Price Comparison 💲
    3. Quit 👋""")
    while True:
        choice = input("Enter your choice (1/2/3): ")
        if choice == '1':
            instructions()
            print()
        elif choice == '2':
            get_items()
            print()
        elif choice == '3':
            print("Thank you for using the Price Comparison Calculator. Goodbye!")
            break
        else:
            print("Invalid choice. Please choose a valid option (1/2/3).")
            continue


# Call the start function to initialize the turtle graphics screen
start()

# Call the show_menu function to display the main menu
show_menu()
